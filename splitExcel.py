# 实现excel表格按行数分拆的功能
import os

import openpyxl


def split_excel():
    # 填入文件名
    excel_name_input = input("填写excel文件名称，限定为.xlsx格式: ")
    excel_name_open = excel_name_input + ".xlsx"

    # 读取对应excel文件，应加入错误提示，待优化
    workbook = openpyxl.load_workbook(filename=excel_name_open)
    sheet_origin = workbook.active  # 获取活跃的表格

    # 获取原表格中限定条数据，并复制到新表格

    rows = sheet_origin.max_row  # 最大行数
    column = sheet_origin.max_column  # 最大列数

    file_path = input("存储文件夹: ")
    # 判断目录是否存在
    if not os.path.exists(file_path):
        # 目录不存在创建，makedir可以创建多级目录
        os.makedirs(file_path)

    # 总共需要多少excel

    limit = int(input("输入分页数据量 "))
    sheets = rows / limit
    if not sheets.is_integer():  # 如果不是整除则需要+1
        sheets = int(sheets) + 1

    for i in range(1, sheets + 1):
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']
        # # 写入第一行数据
        # for n in range(1, column + 1):
        #     sheet.cell(row=1, column=n).value = sheet_origin.cell(row=1, column=n).value
        # 写入范围内数据
        t = 1 + limit * (i - 1)
        num_index = 1
        for row_num in range(t, t + limit ):
            for col_num in range(1, column + 1):
                sheet.cell(row=num_index, column=col_num).value = sheet_origin.cell(row=row_num, column=col_num).value
            num_index = num_index + 1

        wb.save(f'{file_path}/{i}.xlsx')
    print('已完成数据拆分')


if __name__ == '__main__':
    split_excel()
